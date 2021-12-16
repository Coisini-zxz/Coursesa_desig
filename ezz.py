import openpyxl
import json

path = r'C:\Users\Administrator\Desktop\r2.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active
print('activing ---', ws.title)

list1 = []
list2 = []

scope = ws.iter_cols(min_row=2, max_row=222, min_col=2, max_col=2)
for cols in scope:
    for cell in cols:
        list1.append(cell.value)

scope = ws.iter_cols(min_row=2, max_row=222, min_col=3, max_col=3)

for cols in scope:
    for cell in cols:
        list2.append(cell.value)


a = dict(zip(list1, list2))
b = json.dumps(a,  indent=1, ensure_ascii=False)

print(b)
